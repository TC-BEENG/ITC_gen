import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from tkinter import filedialog  # Specifically to select files
import pandas as pd
import datetime
import numpy as np

# For PDFing


saveFilePath = ""

def getTemplate():
    filePath = filedialog.askopenfilename(title = "Select Template Excel File")
    if filePath== "":
        print('No file selected.\n')
        return
    return filePath



def getDirectory():
    folderPath = filedialog.askdirectory()
    if folderPath == "":
        print('No file selected.\n')
        return
    return folderPath

# Get directory to save ITC files.
def getPath(lbl):
    filePath = filedialog.askdirectory(title='Select File Path')
    lbl.configure(text="File Path: {}".format(filePath))  # Update label
    global saveFilePath
    saveFilePath= filePath
    print(saveFilePath)
    return filePath


def makeSheetTest():
    filePath=getTemplate() # Get the file path of the template excel file.
    savePath = getDirectory()
    # Timing- start after file select
    start = datetime.datetime.now()

    # With openpyxl
    wb_template = openpyxl.load_workbook(filePath)
    ws_FAT_template = wb_template["FAT_Template"]

    # With pandas
    df_setup = pd.read_excel(filePath, "Setup")
    df_itcLayout = pd.read_excel(filePath, "ITC_Layout")
    df_FAT_template = pd.read_excel(filePath, "FAT_Template") # TODO: Rename sheet to just template?
    df_SAT_template = pd.read_excel(filePath, "SAT_Template") # Unused for now.

    projectName = df_setup["Project Name"][0]
    revision = df_setup["Revision"][0]

    # Get Each FAT_Template Type in a list
    ITC_types = (df_FAT_template["TYPE"].unique())

    # make dict which is {ITC name: Dataframe of template}
    df_FAT_template_dict = dict.fromkeys(ITC_types, pd.DataFrame()) # initialise empty dataframes

    # For template in ITC_types:
    for idx, r in df_FAT_template.iterrows():
        df_FAT_template_dict[r.iloc[0]] = pd.concat([df_FAT_template_dict[r.iloc[0]], r], axis=1) # construct template dataframes

    # Need to transpose each dataframe to get as desired.
    for key in df_FAT_template_dict.keys():
        df_FAT_template_dict[key] = df_FAT_template_dict[key].transpose()

    # Loop by ITC type so we can make a workbook for each.
    for ITC in ITC_types:
        if ITC == "FAT_HDR" or ITC== "FAT_FTR": # Dont make a sheet for FAT_HDR or footer
            continue
        print("Generating {}...".format(ITC))
        wb = Workbook() # Workbook for each itc type
        sheets_lst = [] # list of generated sheets, to be shown on the cover page.
        desc_lst = [] # With a corresponding description
        for idx, row in df_itcLayout.iterrows():
            # First check if ITC is the one we are looping for
            if not row["ITC Type"] == ITC: # If its not then continue.
                continue
            # 2nd check if generate ITC is yes.
            if not type(row["Generate ITC"]) == str: # If nothing is entered, type is not string
                continue
            if row["Generate ITC"].upper().startswith("Y"):                 # If yes then we need to generate the ITC.
                sheetName = "{}-{:02d}".format(row["ITC DOCUMENT"], row["Sheet No"])
                sheets_lst.append(sheetName)
                desc_lst.append("{}: {}".format(row["DEVICE_ID"], row["DESCRIPTION"]) )
                ws = wb.create_sheet(sheetName)  # Create new worksheet for each row.
                # Copy over dataframe for that ITC, concat with header and footer rows.
                df_temp = pd.concat([df_FAT_template_dict["FAT_HDR"], df_FAT_template_dict[ITC], df_FAT_template_dict["FAT_FTR"]])
                df_temp = df_temp.drop(columns = 'TYPE') # Drop TYPE column

                # Need to make dict for the substitutions {[field from FAT_Template Sheet]: Value from ITC Layout}
                # Perform substitutions per row.
                # Device ID and Description replacing as well
                # sub_dict_keys = list(row.iloc[7:].index) # Gets series of substitution fields.
                sub_dict_keys = list(row.iloc[:].index)  # Gets series of substitution fields.

                sub_dict_keys.extend(["DEVICE_ID", "DESCRIPTION"])
                #sub_dict_keys = ["\[" + x + "\]" for x in sub_dict_keys] # not required anymore since optional col added.

                #sub_dict_vals = list(row.iloc[7:])
                sub_dict_vals = list(row.iloc[:])
                sub_dict_vals.extend([row["DEVICE_ID"], row["DESCRIPTION"]])
                sub_dict_vals2 = [str(x) for x in sub_dict_vals] # Convert all vals to strings, important for replace

                sub_dict = dict(zip(sub_dict_keys, sub_dict_vals2))

                # for k in list(sub_dict.keys()): # remove nans
                #     if sub_dict[k] == "nan":
                #         sub_dict.pop(k)

                # Perform replacement
                df_temp = df_temp.replace(sub_dict, regex=True)

                # Deal with optional rows
                # These will have substitutable fields that have not been substituted.
                dropRows = []
                for idx, r in df_temp.iterrows():  # Row is tuple
                    if "{nan}" == str(r["RowEnable"]): # Searching last columnm the enable column.
                        dropRows.append(idx)
                df_temp = df_temp.drop(dropRows)
                df_temp = df_temp.drop(columns = "RowEnable") # Then drop the enable (last) column.

                # Dataframe to rows.
                for r in dataframe_to_rows(df_temp, index=False, header=False):
                    ws.append(r)

                formatITC(ws)  # Apply print settings to worksheet.

        # Make cover sheet
        df_cover = pd.DataFrame()
        df_cover["Sheet Name"] = sheets_lst
        df_cover["Description"] = desc_lst
        for r in dataframe_to_rows(df_cover, index=False, header=True):
            wb["Sheet"].append(r)
        formatCoverSheet(projectName, wb["Sheet"])

        try:
            wb.save("{}/{}.xlsx".format(savePath, ITC)) # Save workbook.
        except:
            wb.create_sheet("Sheet")
            wb.save("{}/{}.xlsx".format(savePath, ITC)) # Save workbook.
        print("{}.xlsx saved".format(ITC))

    end = datetime.datetime.now()
    #timing
    elapsed = end - start
    print("Make Sheets: {}".format(elapsed.total_seconds()))
    return


def formatITC(ws):
    # Variables that could be changed?
    rows_hdr = 5 # The number of header rows in the template.
    rows_ftr = 9
    rowCount = ws.max_row
    rowCount_str = str(rowCount)
    rows_body = rowCount - (rows_hdr + rows_ftr)
    row_ftr_start = rowCount - rows_ftr + 1

    thin = openpyxl.styles.Side(border_style="thin") # Define thin border style.
    border_all = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
    border_left = openpyxl.styles.Border(left=thin)
    border_right = openpyxl.styles.Border(right=thin)
    border_top= openpyxl.styles.Border(top=thin)
    border_bot = openpyxl.styles.Border(bottom=thin)
    border_horizontal = openpyxl.styles.Border(top=thin, bottom=thin)
    # Corners are different too
    border_tl = openpyxl.styles.Border(top=thin, left=thin)     # top left
    border_tr = openpyxl.styles.Border(top=thin, right=thin)    # bot right etc.
    border_bl = openpyxl.styles.Border(left=thin, bottom=thin)
    border_br = openpyxl.styles.Border(right=thin, bottom=thin)

    ###  Apply outside borders
    # LHS
    range_temp = "A1:A" + rowCount_str
    for row in ws[range_temp]:
        for cell in row:
            cell.border = border_left
    # RHS
    range_temp = "G1:G" + rowCount_str
    for row in ws[range_temp]:
        for cell in row:
            cell.border = border_right
    # Top
    range_temp = "A1:G1"
    for row in ws[range_temp]:
        for cell in row:
            cell.border = border_top
    # Bottom
    range_temp = "A" + rowCount_str + ":G" + rowCount_str
    for row in ws[range_temp]:
        for cell in row:
            cell.border = border_bot
    # and corners
    ws["A1"].border = border_tl
    ws["G1"].border = border_tr
    ws["A" + rowCount_str].border = border_bl
    ws["G" + rowCount_str].border = border_br

    # Column Widths
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 65
    ws.column_dimensions['E'].width = 75
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30

    # Row Heights. Change for body cells to autofit but minimum of 30
    # Autofits by default.
    # We set minumum by changing the font size of an unused column (PASS which is F) to the desired height.
    for row in ws["F8:F{}".format(str(rows_body+rows_hdr))]:
        for cell in row:
            cell.font = openpyxl.styles.Font(size = 30)

    # Did not work.
    #for rowI in range(6, rowCount - rows_ftr + 1): # Loops from start of body to start of footer.
        #ws.row_dimensions[rowI].height = None # None gives us autofit
        #ws.row_dimensions[rowI].auto_size = True  # None gives us autofit
        #ws.row_dimensions[1].height = max(ws.row_dimensions[1].height, 30) not compatible to autofit and

    ### Header formatting
    d2 = ws["D2"] # Bit above title
    d2.font = openpyxl.styles.Font(size = 14, bold = True)
    d2.alignment = openpyxl.styles.Alignment(horizontal = 'center', wrap_text = True)
    ws.merge_cells("D3:D4")
    d3 = ws["D3"]  # Title
    d3.font = openpyxl.styles.Font(size = 18, bold = True)
    d3.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
    F1_5 = ws["F1:F5"]
    for cell in F1_5: # Bold
        cell[0].font = openpyxl.styles.Font(bold = True)
    # Logo
    img = openpyxl.drawing.image.Image('BEE_DarkBlue.png')
    img.width = 165
    img.height = 60
    ws.add_image(img,'B2')
    # Date cell
    ws["G4"] = str(datetime.datetime.today().strftime("%d/%m/%y"))
    # Template Cell
    #ws["G4"] =


    # Body formatting
    ws.merge_cells("A6:G6") # blue bar
    a6 = ws["A6"]
    a6.fill = openpyxl.styles.PatternFill("solid", fgColor='8DB4E2')
    a6.font = openpyxl.styles.Font(bold=True, size=12)
    a6.alignment = openpyxl.styles.Alignment(vertical='center')
    for row in ws["A7:G7"]:  # Bold column headings
        for cell in row:
            cell.font = openpyxl.styles.Font(bold = True)

    bdyRange = "A6:G" + str(rowCount-rows_ftr)
    for row in ws[bdyRange]:  # Bold
        for cell in row:
            cell.border = border_all
            cell.alignment = openpyxl.styles.Alignment(vertical='center', wrap_text=True)

    # Center align test number column
    for row in ws["A8:A{}".format(rowCount-rows_ftr)]:
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    ### Footer Formatting
    # Inspection Sign off row
    ws.merge_cells("A" + str(row_ftr_start) + ":G" + str(row_ftr_start))
    cell_ftr_start = ws["A" + str(row_ftr_start)]
    cell_ftr_start.font = openpyxl.styles.Font(bold = True)
    cell_ftr_start.border = border_all
    for row in ws["A{}:G{}".format(str(row_ftr_start),str(row_ftr_start))]:
        for cell in row:
            cell.border = border_all
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor='D9D9D9')

    # Remarks or remedial action r + str(ows.
    ws.merge_cells("A" + str(row_ftr_start + 4) + ":G" + str(row_ftr_start + 4))
    ws.merge_cells("A" + str(row_ftr_start+5) + ":G" + str(row_ftr_start+5))
    ws.merge_cells("A" + str(row_ftr_start+6) + ":G" + str(row_ftr_start+6))
    ws.merge_cells("A" + str(row_ftr_start+7) + ":G" + str(row_ftr_start+7))
    ws.merge_cells("A" + rowCount_str + ":G" + rowCount_str)

    # Add their border
    for row in ws["A" + str(row_ftr_start + 4) + ":G" + rowCount_str]:
        for cell in row:
            cell.border = border_all
            cell.font = openpyxl.styles.Font(bold = True)

    # Remedial action completed row
    ws.row_dimensions[rowCount].height = 25
    ws["A{}".format(rowCount)].alignment = openpyxl.styles.Alignment(vertical='center')

    ### Print settings
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Page Margins
    ws.page_margins.left = 0.5 # What is the unit? inch?
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.1
    ws.page_margins.bottom = 0.3
    ws.page_margins.footer = 0.2

    # Print area
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    # ws.print_title_cols = "A:G"  # Rows/ cols to repeat when printing over page
    ws.print_title_rows = "7:7"   #
    ws.print_area = ws.dimensions

    # Optimise horizontal page breaks.
    # Adding page break to footer row where it would otherwise run across a page
    # TODO: Make this better
    # if ws.max_row > 29: # 29 just picked ad hoc
    #     pageBreak = openpyxl.worksheet.pagebreak.Break(id=ws.max_row-5)
    #     ws.row_breaks.append(pageBreak)

    # Header and footer of printed sheet
    ws.oddFooter.left.text = ws.title
    ws.oddFooter.right.text = "Page &P"

def formatCoverSheet(projectName, ws):
    thin = openpyxl.styles.Side(border_style="thin")  # Define thin border style.
    border_all = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

    # Add grid to all cells.
    for row in ws[ws.dimensions]:
        for cell in row:
            cell.border = border_all

    # Bold headers.
    for row in ws["A1:B1"]:
        for cell in row:
            cell.font = openpyxl.styles.Font(bold = True)

    # Iterate over all columns and adjust their widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Print settings
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Page Margins
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 1
    ws.page_margins.bottom = 1
    ws.page_margins.header = 0.5


    # Print area
    ws.print_area = ws.dimensions
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.print_options.horizontalCentered = True

    # Header and footer of printed sheet
    ws.oddHeader.left.text = projectName
    ws.oddFooter.left.text = "ITC Index Sheet"
    ws.oddFooter.right.text = "Page &P"

    return


def mainGUI():
    print('Loading Application Window...')
    root = tk.Tk()  # Create root window instance.
    print('Load Complete')
    root.title('ITC Generator')
    root.geometry('400x400')
    root.resizable(False, False)
    btn_width = 30 # in text units
    btn_hght = 2

    label_path = tk.Label(root, text="File Path")
    btn_getpath = tk.Button(root, text='Download Log Files From Webserver',
                           command = lambda: getPath(label_path), width=btn_width, height=btn_hght)

    #btn_makeSheet=

    # Organise
    btn_getpath.grid(row = 0, column = 0, sticky = "w", padx = 2)
    label_path.grid(row =1, column = 0, sticky = "w", padx = 2)
    root.mainloop()
    # rename_btn = tk.Button(root, text='Rename Log Files \n (Append Start Time)',
    #                        command=renameLogs, width=btn_width, height=btn_hght)
    # quit_btn = tk.Button(root, text='Quit',
    #                      command=root.destroy, width=btn_width, height=btn_hght)
    #
    # # Organise (pack) buttons
    # dlLogs_btn.pack(anchor= 'n', padx=10, pady=10)
    # rename_btn.pack(anchor= 'center', padx=10, pady=10)
    # quit_btn.pack(anchor ='s', padx=10, pady=10)
    # root.mainloop()


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    start = datetime.datetime.now()

    #getTemplate()
    makeSheetTest()

    end = datetime.datetime.now()

    elapsed = end-start
    print("Total Time = {}".format(elapsed.total_seconds()))
    #mainGUI()
    #print_hi('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
